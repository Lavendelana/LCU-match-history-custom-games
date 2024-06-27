import openpyxl
from openpyxl.styles import Font
import os.path
import math
from helpers import get_champion_name

from pathlib import Path

class Participant:
  def __init__(self, summonerName, participantData, patch):
    self.summonerName = summonerName
    self.assists=participantData['stats']['assists']
    self.deaths=participantData['stats']['deaths']
    self.firstBloodKill='Yes' if participantData['stats']['firstBloodKill'] == True else 'No'
    self.goldEarned=participantData['stats']['goldEarned']
    self.kills=participantData['stats']['kills']
    self.controlScore=participantData['stats']['timeCCingOthers']
    self.totalDamageDealtToChampions=participantData['stats']['totalDamageDealtToChampions']
    self.creepScore=participantData['stats']['totalMinionsKilled']
    self.visionScore=participantData['stats']['visionScore']
    self.controlWards=participantData['stats']['visionWardsBoughtInGame']
    self.wardsKilled=participantData['stats']['wardsKilled']
    self.wardsPlaced=participantData['stats']['wardsPlaced']
    self.championName=get_champion_name(participantData['championId'], patch)

class GameData:
  def __init__(self, gameData, teamData):
    self.participants = []
    self.baronKills = teamData['baronKills']
    self.dragonKills = teamData['dragonKills']
    self.voidGrubKills = teamData['hordeKills']
    self.inhibitorKills = teamData['inhibitorKills']
    self.riftHeraldKills = teamData['riftHeraldKills']
    self.towerKills = teamData['towerKills']
    self.win = 'Yes' if teamData['win'] == 'Win' else 'No'
    self.gameDate = gameData['gameCreationDate']
    gameDurationTime = gameData['gameDuration']
    self.gameDuration = f'{math.floor(gameDurationTime / 60)}:{gameDurationTime % 60}'
  
  def add_participant(self, participant):
    self.participants.append(participant)

  def write_to_excel(self):
    path = Path('output')
    if not path.exists():
      path.mkdir(parents=True)
    pathToSheet = './output/data.xlsx'
    if os.path.isfile(pathToSheet):
      workbook = openpyxl.load_workbook(pathToSheet)
      worksheet = workbook.active
    else:
      workbook = openpyxl.Workbook()
      worksheet = workbook.active
      # Create the top row after creating a new sheet
      topRow = ["Scrims", "Players", "Kills", "Deaths", "Assists", "Damage dealt to champs", "Firstblood", "Gold", "CS", "Vision Score", "Wards Placed", 
                "Wards Destroyed", "Control Wards", "Crowd Control", "Champion", "Win?", "Towers", "Inhibitors", "Drakes", "Heralds", "Barons", "Voidgrubs", "Match Length"]
      worksheet.append(topRow)
      # Set column widths
      column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(worksheet.max_column))
      for column_letter in column_letters:
        worksheet.column_dimensions[column_letter].bestFit = True
      # Make the top row Bold
      bold_font = Font(color='00000000', bold=True)

      # Enumerate the cells in the first row
      for cell in worksheet["1:1"]:
          cell.font = bold_font
    for index in range(0,5):
      player = self.participants[index]
      newRow = [self.gameDate, player.summonerName, player.kills, player.deaths, player.assists, player.totalDamageDealtToChampions, player.firstBloodKill, player.goldEarned, 
                player.creepScore, player.visionScore, player.wardsPlaced, player.wardsKilled, player.controlWards, player.controlScore, player.championName]
      if (index == 0):
        newRow += [self.win, self.towerKills, self.inhibitorKills, self.dragonKills, self.riftHeraldKills, self.baronKills, self.voidGrubKills, self.gameDuration]
      worksheet.append(newRow)
    
    workbook.save(pathToSheet)
    print('Successfully written to workbook')